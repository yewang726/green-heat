"""
Created on Thu Jan 12 10:21:54 2023 @author: Ye Wang

Get the CF and LCOH2 results for different locations, present-day and future costs, from the HILT Green Hydrogen project  
"""
from greenheatpy.projdirs import hilt_svn_repo
from openpyxl import load_workbook
import numpy as np

def get_data(year, savedir=None):

    fn=hilt_svn_repo+'/H2-SUPPLY/reporting/Final report data/Results with Windlab+Solcast data - %s.xlsx'%year
    wb = load_workbook(fn, data_only=True) 

    locations= wb.sheetnames  
    LCOH2=np.array([])
    LCOH=np.array([])
    title=np.array([])
    PV=np.array([])
    WIND=np.array([])
    EL=np.array([])
    H2ST=np.array([])

    for location in locations:
        ws=wb[location]
        CF=get_row_values(ws, min_row=7, max_row=12, min_col=1, max_col=1) 
        lcoh2=get_row_values(ws, min_row=7, max_row=12, min_col=33, max_col=33) 
        pv=get_row_values(ws, min_row=7, max_row=12, min_col=3, max_col=3) #kW
        wind=get_row_values(ws, min_row=7, max_row=12, min_col=4, max_col=4) #kW   
        electrolyser=get_row_values(ws, min_row=7, max_row=12, min_col=5, max_col=5) #kW H2 (kg)=power(kW) * efficiency/39.4 , efficiency=0.7
        h2_storage=get_row_values(ws, min_row=7, max_row=12, min_col=6, max_col=6) #kg H2, net H2 supply is 5 kg/s

        lcoh=convert_lcoh(lcoh2)
        LCOH2=np.append(LCOH2, lcoh2) #USD/kg
        LCOH=np.append(LCOH, lcoh) # USD/MWh
        title=np.append(title, location)
  
        PV=np.append(PV, pv)
        WIND=np.append(WIND, wind)
        EL=np.append(EL, electrolyser)
        H2ST=np.append(H2ST, h2_storage)
      
    LCOH=LCOH.reshape(len(locations), len(CF))
    LCOH2=LCOH2.reshape(len(locations), len(CF))
    PV=PV.reshape(len(locations), len(CF))
    WIND=WIND.reshape(len(locations), len(CF))
    EL=EL.reshape(len(locations), len(CF))
    H2ST=H2ST.reshape(len(locations), len(CF))
      
    if savedir!=None:
        CF=CF.reshape(len(CF), 1)
        LCOH=np.hstack((CF, LCOH.T))
        LCOH2=np.hstack((CF, LCOH2.T)) 
 
        title1=np.append('CF/LCOH (USD/MWh)', title)
        LCOH=np.vstack((title1, LCOH))
        title2=np.append('CF/LCOH2 (USD/kg)', title)
        LCOH2=np.vstack((title2, LCOH2))  
        data=np.vstack((LCOH, LCOH2))
        np.savetxt(savedir+'/CF_LCOH2_%s.csv'%year, data, delimiter=',', fmt='%s') 

    for i in range(len(locations)):
        locations[i]=locations[i][:-5]
        if 'Pinjara' in locations[i]:
            locations[i]='Pinjarra '+locations[i][-1]

    return CF, LCOH, LCOH2, locations, PV, WIND, EL, H2ST  

def get_storage_data(year, location, cost):

    fn=hilt_svn_repo+'/H2-SUPPLY/reporting/Final report data/Results with Windlab+Solcast data - storage cost impact - %s.xlsx'%year
    wb = load_workbook(fn, data_only=True) 

    title=np.array([])
    sheetnames= wb.sheetnames

    sheetname='%s @%s-%s'%(location, cost, year)

    ws=wb[sheetname]
    CF=get_row_values(ws, min_row=7, max_row=12, min_col=1, max_col=1) 
    LCOH2=get_row_values(ws, min_row=7, max_row=12, min_col=33, max_col=33) 
    LCOH=convert_lcoh(LCOH2)
    pv=get_row_values(ws, min_row=7, max_row=12, min_col=3, max_col=3) #kW
    wind=get_row_values(ws, min_row=7, max_row=12, min_col=4, max_col=4) #kW   
    electrolyser=get_row_values(ws, min_row=7, max_row=12, min_col=5, max_col=5) #kW H2 (kg)=power(kW) * efficiency/39.4 , efficiency=0.7
    h2_storage=get_row_values(ws, min_row=7, max_row=12, min_col=6, max_col=6) #kg H2, net H2 supply is 5 kg/s    

    return CF, LCOH, LCOH2, pv, wind, electrolyser, h2_storage 


def convert_lcoh(lcoh2):
    '''
    Convert levelised cost of hydrogen (LCOH2) in USD/kg to levlised cost of heat (LCOH) in USD/MWh

    Argument:
        lcoh2 (float): LCOH2 in USD/kg
    Return:
        lcoh (float):  LCOH in USD/MWh
    '''
    LHV=120 # MJ/kg
    lcoh=lcoh2/LHV*3600
    return lcoh

def get_row_values(worksheet, min_row, max_row, min_col, max_col):
    row_data=np.array([])
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
             row_data=np.append(row_data, cell.value)
    return row_data

def get_best_location(year, verbose=False):
    '''
    get the best local location in each region that produced the lowest LCOH2 for hydrogen
    '''
    regions=np.array(['Pilbara', 'Pinjarra', 'Gladstone', 'Burnie', 'Upper Spencer Gulf'])
    CF, LCOH, LCOH2, locations, PV, WIND, EL, H2ST =get_data(year, savedir=None)
    best={}
    for region in regions:
        data=np.array([])
        for i in range(len(locations)):
            location=locations[i]
            if region in location:
                data=np.append(data, LCOH2[i,-1])
        best_id=np.argmin(data)+1
        if verbose:
            print(region, 'has total locations: ', len(data), ', the best location is No.', best_id)
        best[region]=best_id

    return best
       
def plot_bar(locations, LCOH2):
    import matplotlib.pyplot as plt
    fts=14
    fig, ax = plt.subplots()
    ax.bar(locations, LCOH2, width=0.4)
    ax.tick_params(axis='x', labelrotation=45, labelsize=fts)
    ax.tick_params(axis='y', labelsize=fts)
    #plt.xlabel(fontsize=20)
    plt.ylabel('LCOH2 (USD/kg)', fontsize=fts)
    plt.savefig(open('./LCOH2.png', 'wb'), bbox_inches='tight', dpi=200)
    #plt.show()
    plt.close()            

if __name__=='__main__':

    years=[2020, 2030, 2050]
    for year in years:
        CF, LCOH, LCOH2, locations, PV, WIND, EL, H2ST =get_data(year, savedir='./')
        #best=get_best_location(year, verbose=True)
        #plot_bar(locations, LCOH2[:,-2])
        #print(best)
        #CF, LCOH, LCOH2=get_storage_data(year=2020, location='Pilbara 3', cost=1000)
        #print(CF, LCOH, LCOH2)
        stop
